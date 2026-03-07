from collections import Counter
from io import BytesIO
import json
from tempfile import NamedTemporaryFile
from django import forms
from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.urls import path
from django.utils.encoding import escape_uri_path
from django.utils.safestring import mark_safe
import openpyxl
from .models import (
    Chapter,
    ChoiceProblem,
    CodeDraft,
    Course,
    DiscussionThread,
    Problem,
    AlgorithmProblem,
    ProblemProgress,
    ProblemUnlockCondition,
    Submission,
    TestCase,
    FillBlankProblem,
    Exam,
    ExamProblem,
    ExamSubmission,
    ExamAnswer,
    Enrollment,
    ChapterUnlockCondition,
    ChapterProgress,
    CourseUnlockSnapshot,
    ProblemUnlockSnapshot,
)
from .course_import_services.git_repo_service import GitRepoService
from .course_import_services.course_importer import CourseImporter
# Register your models here.


class ImportCourseFromRepoForm(forms.Form):
    repo_url = forms.CharField(
        label="Git Repository URL",
        widget=forms.TextInput(attrs={"size": 60}),
        help_text="HTTPS or SSH URL to Git repository",
    )
    branch = forms.CharField(
        label="Branch",
        initial="main",
        required=False,
        help_text="Git branch to clone (default: main)",
    )
    update_mode = forms.BooleanField(
        label="Update existing courses",
        initial=True,
        required=False,
        help_text="If checked, update existing courses; otherwise skip them",
    )


@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ("title", "created_at", "updated_at")
    search_fields = ("title", "description")
    change_list_template = "admin/course_change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-from-repo/",
                self.admin_site.admin_view(self.import_from_repo_view),
                name="course_import_from_repo",
            ),
        ]
        return custom_urls + urls

    def import_from_repo_view(self, request):
        if request.method == "POST":
            form = ImportCourseFromRepoForm(request.POST)
            if form.is_valid():
                repo_url = form.cleaned_data["repo_url"]
                branch = form.cleaned_data.get("branch") or "main"
                update_mode = form.cleaned_data.get("update_mode", True)

                try:
                    with GitRepoService(repo_url, branch) as repo_path:
                        importer = CourseImporter(repo_path, update_mode=update_mode)
                        stats = importer.import_all()

                        # Format results
                        results = f"""
                        <h2>Import Summary</h2>
                        <ul>
                            <li>Courses created: <strong>{stats["courses_created"]}</strong></li>
                            <li>Courses updated: <strong>{stats["courses_updated"]}</strong></li>
                            <li>Courses skipped: <strong>{stats["courses_skipped"]}</strong></li>
                            <li>Chapters created: <strong>{stats["chapters_created"]}</strong></li>
                            <li>Chapters updated: <strong>{stats["chapters_updated"]}</strong></li>
                            <li>Problems created: <strong>{stats["problems_created"]}</strong></li>
                            <li>Problems updated: <strong>{stats["problems_updated"]}</strong></li>
                        </ul>
                        """

                        if stats["errors"]:
                            results += f"<h3>Errors ({len(stats['errors'])})</h3><ul>"
                            for error in stats["errors"][:10]:
                                results += f"<li>{error}</li>"
                            if len(stats["errors"]) > 10:
                                results += f"<li>... and {len(stats['errors']) - 10} more errors</li>"
                            results += "</ul>"

                        self.message_user(
                            request, "Course import completed!", level="success"
                        )

                        return render(
                            request,
                            "admin/course_import_results.html",
                            {
                                "results": results,
                                "opts": self.model._meta,
                            },
                        )

                except Exception as e:
                    self.message_user(
                        request, f"Import failed: {str(e)}", level="error"
                    )
        else:
            form = ImportCourseFromRepoForm()

        context = dict(
            self.admin_site.each_context(request),
            title="Import Courses from Git Repository",
            form=form,
            opts=self.model._meta,
        )
        return render(request, "admin/course_import_from_repo.html", context)


# admin.site.register(Chapter)
# admin.site.register(Problem)
admin.site.register(AlgorithmProblem)
admin.site.register(ChoiceProblem)
admin.site.register(FillBlankProblem)
admin.site.register(Submission)
admin.site.register(TestCase)
admin.site.register(ProblemProgress)
admin.site.register(DiscussionThread)
admin.site.register(ProblemUnlockCondition)
admin.site.register(CodeDraft)
admin.site.register(Exam)
admin.site.register(Enrollment)
admin.site.register(ChapterProgress)


class ImportProblemForm(forms.Form):
    chapter = forms.ModelChoiceField(
        queryset=Chapter.objects.all(),
        required=False,
        label="所属章节（可选）",
        help_text="若选择，所有导入题目将关联到该章节",
    )
    excel_file = forms.FileField(label="上传 Excel 文件")


class ImportChapterForm(forms.Form):
    course = forms.ModelChoiceField(
        queryset=Course.objects.all(), required=True, label="所属课程"
    )
    excel_file = forms.FileField(
        label="上传 Excel 文件", help_text="文件需包含 'chapters' 工作表"
    )


@admin.register(Problem)
class ProblemAdmin(admin.ModelAdmin):
    list_display = ("title", "type", "chapter", "difficulty", "created_at")
    list_filter = ("type", "chapter", "difficulty")
    search_fields = ("title",)
    change_list_template = "admin/problem_change_list.html"
    actions = ["export_problems_to_excel"]  # ← 注册 action

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_view),
                name="problem_import",
            ),
            path(
                "import/template/",
                self.admin_site.admin_view(self.download_template),
                name="problem_import_template",
            ),
        ]
        return custom_urls + urls

    def import_view(self, request):
        if request.method == "POST":
            form = ImportProblemForm(request.POST, request.FILES)
            if form.is_valid():
                chapter = form.cleaned_data.get("chapter")
                excel_file = request.FILES.get("excel_file")
                if not excel_file:
                    form.add_error("excel_file", "请上传 Excel 文件")
                else:
                    try:
                        self._handle_import(excel_file, chapter)
                        self.message_user(
                            request, "✅ 题目已成功导入！", level="success"
                        )
                        # 成功后重定向
                        return redirect(
                            reverse(
                                "admin:%s_%s_changelist"
                                % (
                                    self.model._meta.app_label,
                                    self.model._meta.model_name,
                                )
                            )
                        )
                    except Exception as e:
                        self.message_user(
                            request, f"❌ 导入失败: {str(e)}", level="error"
                        )
            # 如果表单无效或导入失败，继续渲染表单（带错误）
        else:
            form = ImportProblemForm()

        context = dict(
            self.admin_site.each_context(request),
            title="批量导入题目",
            form=form,
            opts=self.model._meta,
        )
        return render(request, "admin/problem_import.html", context)

    def download_template(self, request):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "problems"

        # 表头
        headers = [
            "type",
            "title",
            "content",
            "difficulty",
            "options",
            "correct_answer",
            "is_multiple_choice",
            "time_limit",
            "memory_limit",
            "code_template",
            "solution_name",
        ]
        ws.append(headers)

        # 示例行（选择题）
        ws.append(
            [
                "choice",
                "示例选择题",
                "题目描述",
                2,
                '{"A": "选项A", "B": "选项B"}',
                '"A"',
                "false",
                "",
                "",
                "",
                "",
            ]
        )

        # 示例行（算法题）
        ws.append(
            [
                "algorithm",
                "示例算法题",
                "输入两个整数，输出和",
                1,
                "",
                "",
                "",
                1000,
                256,
                '{"python": "def solve(a, b):\\n    return a + b"}',
                '{"python": "solve"}',
            ]
        )

        # 创建 test_cases 工作表（可选）
        ws2 = wb.create_sheet(title="test_cases")
        ws2.append(["problem_title", "input_data", "expected_output", "is_sample"])
        ws2.append(["示例算法题", "1 2", "3", "true"])
        tmpfile = NamedTemporaryFile()
        wb.save(tmpfile.name)
        response = HttpResponse(
            tmpfile,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            "attachment; filename=problem_import_template.xlsx"
        )
        return response

    def _handle_import(self, excel_file, chapter):
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(filename=BytesIO(excel_file.read()))
        if "problems" not in wb.sheetnames:
            raise ValueError("Excel 文件中缺少 'problems' 工作表")
        ws = wb["problems"]

        problems_data = []
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row[0]:  # 跳过空行
                continue
            # 补齐缺失列（防止列数不足）
            row = (row + (None,) * 11)[:11]  # 确保至少11列
            data = dict(
                zip(
                    [
                        "type",
                        "title",
                        "content",
                        "difficulty",
                        "options",
                        "correct_answer",
                        "is_multiple_choice",
                        "time_limit",
                        "memory_limit",
                        "code_template",
                        "solution_name",
                    ],
                    row,
                )
            )
            problems_data.append(data)
        # 🔍 检查本次导入中 title 是否重复
        titles = [d["title"] for d in problems_data if d.get("title")]
        title_counts = Counter(titles)
        duplicate_titles = [title for title, count in title_counts.items() if count > 1]

        if duplicate_titles:
            dup_list = ", ".join(f'"{t}"' for t in duplicate_titles[:5])  # 最多显示5个
            raise ValueError(
                f"Excel 中存在重复的题目标题，请确保每道题标题唯一。重复标题示例：{dup_list}"
            )
        # 先创建所有 Problem
        problem_objs = []
        for data in problems_data:
            prob = Problem(
                type=data["type"],
                title=data["title"],
                content=data["content"],
                difficulty=int(data["difficulty"]) if data["difficulty"] else 1,
                chapter=chapter,
            )
            prob.full_clean()
            problem_objs.append(prob)
        created_problems = Problem.objects.bulk_create(problem_objs)

        # 按顺序映射（假设标题唯一）
        title_to_problem = {p.title: p for p in created_problems}

        # 创建子表
        algo_list = []
        choice_list = []
        test_cases_list = []

        for data in problems_data:
            prob = title_to_problem.get(data["title"])
            if not prob:
                continue

            if prob.type == "algorithm":
                algo = AlgorithmProblem(
                    problem=prob,
                    time_limit=int(data["time_limit"]) if data["time_limit"] else 1000,
                    memory_limit=int(data["memory_limit"])
                    if data["memory_limit"]
                    else 256,
                    code_template=json.loads(data["code_template"])
                    if data["code_template"]
                    else None,
                    solution_name=json.loads(data["solution_name"])
                    if data["solution_name"]
                    else None,
                )
                algo_list.append(algo)
            elif prob.type == "choice":
                choice = ChoiceProblem(
                    problem=prob,
                    options=json.loads(data["options"]) if data["options"] else {},
                    correct_answer=json.loads(data["correct_answer"])
                    if data["correct_answer"]
                    else [],
                    is_multiple_choice=bool(data["is_multiple_choice"]),
                )
                choice_list.append(choice)

        AlgorithmProblem.objects.bulk_create(algo_list)
        ChoiceProblem.objects.bulk_create(choice_list)

        # 处理测试用例（如果存在 test_cases 表）
        if "test_cases" in wb.sheetnames:
            ws_tc = wb["test_cases"]
            tc_rows = list(ws_tc.iter_rows(min_row=2, values_only=True))
            for row in tc_rows:
                title, inp, out, is_sample = (row + (None,) * 4)[:4]
                prob = title_to_problem.get(title)
                if prob and hasattr(prob, "algorithm_info"):
                    tc = TestCase(
                        problem=prob.algorithm_info,
                        input_data=inp or "",
                        expected_output=out or "",
                        is_sample=bool(is_sample),
                    )
                    test_cases_list.append(tc)
            TestCase.objects.bulk_create(test_cases_list)

    def export_problems_to_excel(self, request, queryset):
        """
        自定义 Action：导出选中的题目为 Excel
        """
        import openpyxl
        from io import BytesIO
        import json

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws_problems = wb.active
        ws_problems.title = "problems"

        # problems 表头
        problem_headers = [
            "type",
            "title",
            "content",
            "difficulty",
            "options",
            "correct_answer",
            "is_multiple_choice",
            "time_limit",
            "memory_limit",
            "code_template",
            "solution_name",
        ]
        ws_problems.append(problem_headers)

        # 存储测试用例
        test_cases_data = []

        for prob in queryset.select_related("chapter").prefetch_related(
            "choice_info", "algorithm_info__test_cases"
        ):
            if prob.type == "choice":
                choice = getattr(prob, "choice_info", None)
                options = (
                    json.dumps(choice.options) if choice and choice.options else ""
                )
                correct_answer = (
                    json.dumps(choice.correct_answer)
                    if choice and choice.correct_answer
                    else ""
                )
                is_mc = choice.is_multiple_choice if choice else False
                time_limit = memory_limit = code_template = solution_name = ""
            elif prob.type == "algorithm":
                algo = getattr(prob, "algorithm_info", None)
                options = correct_answer = is_mc = ""
                time_limit = algo.time_limit if algo else 1000
                memory_limit = algo.memory_limit if algo else 256
                code_template = (
                    json.dumps(algo.code_template)
                    if algo and algo.code_template
                    else ""
                )
                solution_name = (
                    json.dumps(algo.solution_name)
                    if algo and algo.solution_name
                    else ""
                )

                # 收集测试用例
                if algo:
                    for tc in algo.test_cases.all():
                        test_cases_data.append(
                            [
                                prob.title,
                                tc.input_data,
                                tc.expected_output,
                                tc.is_sample,
                            ]
                        )
            else:
                options = correct_answer = is_mc = ""
                time_limit = memory_limit = code_template = solution_name = ""

            ws_problems.append(
                [
                    prob.type,
                    prob.title,
                    prob.content,
                    prob.difficulty,
                    options,
                    correct_answer,
                    is_mc,
                    time_limit,
                    memory_limit,
                    code_template,
                    solution_name,
                ]
            )

        # 如果有测试用例，创建 test_cases 工作表
        if test_cases_data:
            ws_test_cases = wb.create_sheet(title="test_cases")
            ws_test_cases.append(["problem_title", "input", "output", "is_sample"])
            for row in test_cases_data:
                ws_test_cases.append(row)
        tmpfile = NamedTemporaryFile()
        wb.save(tmpfile.name)
        # 生成响应
        response = HttpResponse(
            content=tmpfile,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = "exported_problems.xlsx"
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{escape_uri_path(filename)}"
        )
        return response

    export_problems_to_excel.short_description = "导出选中题目为 Excel"


class ChapterUnlockConditionInline(admin.TabularInline):
    """章节解锁条件内联编辑"""

    model = ChapterUnlockCondition
    can_delete = True
    extra = 0
    verbose_name = "解锁条件"
    verbose_name_plural = "解锁条件"

    fields = (
        "unlock_condition_type",
        "prerequisite_chapters",
        "unlock_date",
    )


@admin.register(Chapter)
class ChapterAdmin(admin.ModelAdmin):
    list_display = ("course", "title", "order", "show_dependent_chapters", "created_at")
    list_filter = ("course",)
    actions = ["export_chapters_to_excel"]
    inlines = [ChapterUnlockConditionInline]
    change_list_template = "admin/chapter_change_list.html"  # 自定义模板加导入按钮

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_view),
                name="chapter_import",
            ),
            path(
                "import/template/",
                self.admin_site.admin_view(self.download_template),
                name="chapter_import_template",
            ),
        ]
        return custom_urls + urls

    def download_template(self, request):
        """提供章节导入模板下载"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "chapters"

        # 表头
        headers = ["title", "content", "order"]
        ws.append(headers)

        # 可选：添加一行示例（灰色注释风格，用户可删）
        example = ["变量与数据类型", "介绍变量、整数、字符串等基础类型", 1]
        ws.append(example)
        tmpfile = NamedTemporaryFile()
        wb.save(tmpfile.name)
        # 设置响应
        response = HttpResponse(
            content=tmpfile,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = "chapter_import_template.xlsx"
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{escape_uri_path(filename)}"
        )
        return response

    def import_view(self, request):
        if request.method == "POST":
            form = ImportChapterForm(request.POST, request.FILES)
            if form.is_valid():
                course = form.cleaned_data.get("course")
                excel_file = request.FILES["excel_file"]
                try:
                    self._handle_import(excel_file, course)
                    self.message_user(request, "章节导入成功！", level="success")
                    return redirect("..")
                except Exception as e:
                    self.message_user(request, f"导入失败: {str(e)}", level="error")
            else:
                self.message_user(request, "请上传有效的 Excel 文件。", level="error")
        else:
            form = ImportChapterForm()
        context = dict(
            self.admin_site.each_context(request),
            title="批量导入章节",
            form=form,
            opts=self.model._meta,
        )
        return render(request, "admin/chapter_import.html", context)

    def _handle_import(self, excel_file, course):
        wb = openpyxl.load_workbook(filename=BytesIO(excel_file.read()))
        if "chapters" not in wb.sheetnames:
            raise ValueError("Excel 文件中缺少 'chapters' 工作表")

        ws = wb["chapters"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        chapters_data = []

        for row in rows:
            if not row or not row[0]:
                continue
            # 补齐列
            row = (row + (None,) * 3)[:3]
            title, content, order = row
            if not title:
                continue
            order = int(order) if order is not None else 0
            chapters_data.append(
                {"title": title.strip(), "content": content or "", "order": order}
            )

        # 2. 检查本次导入中是否有重复的 (course_title, order)
        from collections import Counter

        key_counts = Counter((item["order"]) for item in chapters_data)
        duplicates = [key for key, cnt in key_counts.items() if cnt > 1]
        if duplicates:
            dup_str = ", ".join([f"({c}, order={o})" for c, o in duplicates[:5]])
            raise ValueError(f"章节顺序冲突（同一课程内 order 重复）: {dup_str}")

        # 3. 创建章节对象
        chapter_objs = []
        for item in chapters_data:
            course = course
            chapter = Chapter(
                course=course,
                title=item["title"],
                content=item["content"],
                order=item["order"],
            )
            chapter.full_clean()
            chapter_objs.append(chapter)

        # 4. 批量创建
        Chapter.objects.bulk_create(chapter_objs)

    # ========================
    # 导出 Action
    # ========================
    def export_chapters_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "chapters"
        ws.append(["course_title", "title", "content", "order"])

        for chapter in queryset.select_related("course"):
            ws.append(
                [chapter.course.title, chapter.title, chapter.content, chapter.order]
            )
        tmpfile = NamedTemporaryFile()
        wb.save(tmpfile.name)
        response = HttpResponse(
            content=tmpfile,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = "exported_chapters.xlsx"
        response["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{escape_uri_path(filename)}"
        )
        return response

    export_chapters_to_excel.short_description = "导出选中章节为 Excel"

    def show_dependent_chapters(self, obj):
        """
        显示依赖于当前章节的章节列表
        """
        # 获取所有依赖此章节的章节（即需要完成此章节才能解锁的章节）
        dependent_chapters = (
            Chapter.objects.filter(unlock_condition__prerequisite_chapters=obj)
            .select_related("course")
            .order_by("course__title", "order")
        )

        if dependent_chapters.exists():
            chapter_links = []
            for chapter in dependent_chapters:
                chapter_links.append(
                    f'<a href="/admin/courses/chapter/{chapter.id}/change/" target="_blank">{chapter.course.title} - {chapter.title}</a>'
                )
            return mark_safe("<br>".join(chapter_links))
        else:
            return "无依赖此章节的章节"

    show_dependent_chapters.short_description = "依赖章节"


# ============================================================================
# 测验功能相关 Admin 配置
# ============================================================================


@admin.register(ExamProblem)
class ExamProblemAdmin(admin.ModelAdmin):
    """测验题目管理"""

    list_display = ("exam", "problem", "score", "order", "is_required")
    list_filter = ("exam", "problem__type")
    search_fields = ("exam__title", "problem__title")
    ordering = ("exam", "order")


@admin.register(ExamSubmission)
class ExamSubmissionAdmin(admin.ModelAdmin):
    """测验提交管理"""

    list_display = (
        "user",
        "exam",
        "status",
        "total_score",
        "is_passed",
        "started_at",
        "submitted_at",
    )
    list_filter = ("exam", "status", "is_passed")
    search_fields = ("user__username", "exam__title")
    ordering = ("-started_at",)


@admin.register(ExamAnswer)
class ExamAnswerAdmin(admin.ModelAdmin):
    """测验答案管理"""

    list_display = ("submission", "problem", "score", "is_correct")
    list_filter = ("problem__type", "is_correct")
    search_fields = ("submission__user__username", "problem__title")
    ordering = ("submission", "problem__id")


# ============================================================================
# 快照功能相关 Admin 配置
# ============================================================================


@admin.register(CourseUnlockSnapshot)
class CourseUnlockSnapshotAdmin(admin.ModelAdmin):
    """课程解锁状态快照管理"""

    list_display = ("course", "enrollment", "is_stale", "version", "computed_at")
    list_filter = ("is_stale", "course", "computed_at")
    search_fields = ("course__title", "enrollment__user__username")
    ordering = ("-computed_at",)
    readonly_fields = (
        "course",
        "enrollment",
        "unlock_states",
        "computed_at",
        "version",
    )


@admin.register(ProblemUnlockSnapshot)
class ProblemUnlockSnapshotAdmin(admin.ModelAdmin):
    """问题解锁状态快照管理"""

    list_display = ("course", "enrollment", "is_stale", "version", "computed_at")
    list_filter = ("is_stale", "course", "computed_at")
    search_fields = ("course__title", "enrollment__user__username")
    ordering = ("-computed_at",)
    readonly_fields = (
        "course",
        "enrollment",
        "unlock_states",
        "computed_at",
        "version",
    )
